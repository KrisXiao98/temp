# -*- coding = utf-8 -*-
# @time:2022/4/18 22:08
# Author:KrisXiao
# @File:excel_conf.py
# @Software:PyCharm

'''
    Excel的配置文件,当前只是添加了增加pass和failed时情况
'''

from openpyxl.styles import PatternFill, Font


# Pass配置
def passed(cell, row, column):
    cell(row=row, column=column).value = 'Pass'
    # 写入单元格样式设定：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
    cell(row=row, column=column).font = Font(bold=True)


# Failed配置
def failed(cell, row, column):
    cell(row=row, column=column).value = 'Failed'
    # 写入单元格样式设定：绿色+加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
    cell(row=row, column=column).font = Font(bold=True)
